import openpyxl

## sets the path to the exel file and the row and colmn variables.
path = "./DataHolder.xlsx"
row=int(1)
column=int(8)
found = {}

## opens the exel file
wb_obj = openpyxl.load_workbook(path)

## sets the sheet
sheet_obj = wb_obj.active

## sets the cell variable
cell_obj = sheet_obj.cell(row=1, column=1)

## Searches for a specific Name, type, or pokemon
def search_data(word):

    
    found = {}
    for cell_obj in sheet_obj:

        if word in cell_obj:

            found[word]=cell_obj

    return found


def Looking_Around(row):
    row = int(row)
    ## Tells you how many you asked for, reaffirming it.
    print("Total Rows:", row) 
    print("Total Columns:", column) 

    ## Goes through each Gym leader till the one selected.
    print("\nValue of first column") 
    for i in range(1, row + 1): 
        cell_obj = sheet_obj.cell(row=i, column=1) 
        print(i, cell_obj.value) 

    ## Goes through the Leader, their type speciality, and the pokemon on the team.  
    print("\nValue of first row") 
    for i in range(1, column + 1): 
        cell_obj = sheet_obj.cell(row, column=i) 
        print(cell_obj.value, end=" ") 
## Asks if you want to look for a specific word or to just search
answer = input("Do you have a specific Leader, type, or pokemon in mind?(YES, NO):")
    ## If yes
if answer == "YES":
   word = input("What are you looking for?(Gym leader, type, or pokemon):")
   search_data(word)
   print(found)

   ## If no
elif answer == "NO":
     ## asks for the leader and team
     row = input("Enter your row: ")
     Looking_Around(row)

else:
    print("That is not a response.")

## hard sets the input to a number variable     



